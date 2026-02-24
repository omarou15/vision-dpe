import pandas as pd
import json
import os
import copy
import yaml
import numpy as np
from pathlib import Path
from lxml.etree import SubElement
from lxml.etree import ElementTree, XMLSchema, Element
from lxml import etree
import openpyxl

WRITE_IN_PLACE = True

## load excel data


def get_sources(el):
    appinfo=el.xpath('.//xs:appinfo', namespaces=namespaces)
    if len(appinfo)>0:
        return appinfo[0].attrib.get('source')
    else:
        return None

path = Path("..") / "modele_donnee"

df_mdd = pd.read_excel(path / 'modele_donnee.xlsx')
df_mdd = df_mdd[['description', 'nom_variable', 'parent_xpath', 'type', 'condition', 'DPE NEUF']]

enum_table = pd.read_excel(path / 'enum_tables.xlsx', sheet_name=None, dtype=str)
print(len(enum_table))
enum_table_audit = pd.read_excel(path / 'enum_tables_audit.xlsx', sheet_name=None, dtype=str)

enum_table_audit['version_audit']['id'] = (enum_table_audit['version_audit'].id.astype(float).round(1)).astype(str)

enum_doc_audit = pd.read_excel(path / 'enum_doc_audit.xlsx', sheet_name='doc', dtype=str)

valeur_table = pd.read_excel(path / 'valeur_tables.xlsx', sheet_name=None)
valeur_table['coef_masque_lointain_non_homogene'] = valeur_table['coef_masque_lointain_non_homoge']
valeur_table['reseau_chaleur'] = valeur_table['reseau_chaleur_2020']
del valeur_table['reseau_chaleur_2020']
del valeur_table['reseau_chaleur_2021']
del valeur_table['coef_masque_lointain_non_homoge']

enums = [el for el in df_mdd.nom_variable.dropna().values.tolist() if el.startswith('enum')]
enums_without_id = [el for el in enums if not el.endswith('id')]

df_mdd['parent_xpath_neuf'] = df_mdd['parent_xpath'].str.replace('dpe.logement.', 'dpe.logement_neuf.')

df_mdd['condition'] = (df_mdd.condition == 'requis').replace({True: 'requis', False: 'optionnel'})

null = (df_mdd['DPE NEUF'] == 'non demandé') | (df_mdd['DPE NEUF'].isnull())
df_mdd.loc[null, 'parent_xpath_neuf'] = np.nan

mdd = df_mdd.dropna(subset=['nom_variable']).to_dict(orient='records')

mdd_neuf = [el.copy() for el in mdd if el['parent_xpath_neuf'] is not np.nan]

for el in mdd_neuf:
    el['parent_xpath'] = el['parent_xpath_neuf']
    condition = el['DPE NEUF']
    if condition != "requis":
        condition = 'optionnel'
    el['condition'] = condition

mdd.extend(mdd_neuf)

# suppresion de la ref dpe/audit pour logement et administratif (qui est maintenant une ref.)
for el in mdd:
    el['parent_xpath'] = el['parent_xpath'].replace('dpe.logement.', 'logement.')
    el['parent_xpath'] = el['parent_xpath'].replace('dpe.administratif', 'administratif')

## write enums dict


enums_dict = {k: v.astype(str).set_index('id').lib.str.strip().str.lower().to_dict() for k, v in enum_table.items() if 'lib' in v and 'id' in v}
enums_dict['classe_etiquette'] = {k: v.upper() for k, v in enums_dict['classe_etiquette'].items()}
for k, v in enums_dict.items():
    for k1, v1 in v.items():
        if isinstance(v1, str):
            v[k1] = v1.replace('\xa0', ' ')

with open(path / 'enums.json', 'w', encoding='utf-8') as f:
    json.dump(enums_dict, f, indent=4, ensure_ascii=False)

enums_audit_doc = enum_doc_audit.set_index('nom enum').to_dict()['documentation']

enums_audit_dict = {k: v[~v.lib.isna()].astype(str).set_index('id').lib.str.strip().str.lower().to_dict() for k, v in enum_table_audit.items() if 'lib' in v and 'id' in v}

# enregistrement des enums audit dans un .json
for enum_name, id_lib in enums_audit_dict.items():
    for _id, lib in id_lib.items():
        if isinstance(lib, str):
            id_lib[_id] = lib.replace('\xa0', ' ')

with open(path / 'enums_audit.json', 'w', encoding='utf-8') as f:
    json.dump(enums_audit_dict, f, indent=4, ensure_ascii=False)

##  load xsd

# ## load xsd

from lxml.etree import ElementTree, XMLSchema
from lxml import etree

parser = etree.XMLParser(remove_blank_text=True)
path_xsd = r'../modele_donnee/modele_commun_DPE_audit.xsd'
# path_xsd = r'.\DPE_dev.xsd'
schema = XMLSchema(file=path_xsd)

parser = etree.XMLParser(remove_blank_text=True)

et = etree.parse(path_xsd, parser)
root = et.getroot()

xs = '{http://www.w3.org/2001/XMLSchema}'
obsdpe = '{https://gitlab.com/observatoire-dpe/observatoire-dpe/-/tree/master/modele_donnee}'

namespaces = {'xs': 'http://www.w3.org/2001/XMLSchema',
              'obsdpe': "https://gitlab.com/observatoire-dpe/observatoire-dpe/-/tree/master/modele_donnee"}

all_el = list(root.iterfind('*//xs:element', namespaces=namespaces))

all_doc = list(root.iterfind('*//xs:documentation', namespaces=namespaces))

## xsd elements ->  object model (dict)

list_keys = list()
list_var = list()
for i, el in enumerate(all_el):
    #     if el.attrib.get('name')=='annee_construction':
    #         qdsqds
    caracs = dict(el.attrib)
    caracs['self'] = el
    if 'name' in caracs:
        is_complex = len([child for child in el.getchildren() if 'complexType' in str(child.tag)]) > 0
        if not is_complex:
            doc = el.find(f'xs:annotation/xs:documentation', namespaces=namespaces)
            if doc is not None:
                caracs.update({'documentation': doc.text})
            appinfo = el.find(f'xs:annotation/xs:appinfo', namespaces=namespaces)
            if appinfo is not None:
                caracs.update(appinfo.attrib)
                if appinfo.text is not None:
                    caracs.update({'appinfo': appinfo.text})
            restriction = el.find(f'xs:simpleType/xs:restriction', namespaces=namespaces)
            if restriction is not None:
                caracs.update(restriction.attrib)
                if len(restriction.getchildren()) >= 2 and restriction.attrib.get('base') == 'xs:int':
                    caracs.update({'len_enum': np.max([int(el.attrib['value']) for el in restriction.getchildren() if not 'pattern' in str(el.tag)])})
            list_var.append(caracs)
for el in list_var:
    if el.get('minOccurs') == '0':
        el['condition'] = 'optionnel'
    else:
        el['condition'] = 'requis'

dpe_neuf = [el for el in list_var if 'logement_neuf' in el.get('source', '')]

for el in list_var:
    el['xpath'] = el.get('source', '')

for el in mdd:
    el['xpath'] = el['parent_xpath'].replace('.', '/') + '/' + el["nom_variable"]

## update version

# # load and update version

import yaml

with open('../versions.yml', 'r', encoding='utf-8') as f:
    versions = yaml.safe_load(f)

v_audit = versions['model_audit']
v_dpe = versions['model_dpe']
v_commun = versions['modele_commun']

model_version_commun = f"Version {v_commun['version']} - {v_commun['date']} : {v_commun['titre']}"

model_version_audit = f"Version V{v_audit['version']} - {v_audit['date']} : {v_audit['titre']}"

model_version_dpe = f"Version V{v_dpe['version']} - {v_dpe['date']} : {v_dpe['titre']}"

dpe = root.xpath(f'//xs:element[@name="dpe"]', namespaces=namespaces)[0]
doc = dpe.xpath('xs:annotation/xs:documentation', namespaces=namespaces)[0]
doc.text = model_version_dpe

audit = root.xpath(f'//xs:element[@name="audit"]', namespaces=namespaces)[0]

documentation = audit.find('xs:annotation/xs:documentation', namespaces=namespaces)
documentation.text = f'{model_version_audit} ////// \n\t En cohérence avec la version dpe : {model_version_dpe}'

documentation = root.find('xs:annotation/xs:documentation', namespaces=namespaces)
documentation.text = f'{model_version_commun} ////// \n\t  {model_version_dpe} ////// \n\t  {model_version_audit}'

## process doc

### reformat doc

all_doc = list(root.iterfind(f'*//xs:documentation', namespaces=namespaces))

for doc in all_doc:
    txt = doc.text.replace('\n', '').strip()
    txt = ' '.join(txt.split())
    doc.text = txt

et.write('test.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

all_el = list(root.iterfind('*//xs:element', namespaces=namespaces))

all_doc = list(root.iterfind('*//xs:documentation', namespaces=namespaces))

for doc in all_doc:
    txt = doc.text.replace('\n', '').strip()
    doc.text = txt

for el in list_var:

    eq_var = [x for x in mdd if el['xpath'] == x['xpath']]

    if len(eq_var) >= 1:
        doc = ' '.join(eq_var[0]['description'].split()).strip()
        self = el['self']
        documentation = self.find(f'xs:annotation/xs:documentation', namespaces=namespaces)
        if documentation is not None:
            documentation.text = doc
        else:
            annotation = self.find(f'xs:annotation', namespaces=namespaces)
            documentation = SubElement(annotation, f'{xs}documentation')
            documentation.text = doc
    else:

        print('not directly found', el['name'])

## process enums

# ## enums
enums_dict.update(enums_audit_dict)

enums_mdd = [el['nom_variable'] for el in mdd if el['nom_variable'].startswith('enum_')]

enums_xsd_name = [el['name'] for el in list_var if el['name'].startswith('enum_')]

enums_xsd = [el for el in list_var if el['name'].startswith('enum_')]

enums_xsd_name_dict = dict()
for el in list_var:
    if el['name'].startswith('enum_'):
        if len(el.get('appinfo', '')) > 0:
            try:
                enums_xsd_name_dict[el['name'].replace('enum_', '').replace('_id', '')] = json.loads(el['appinfo'].replace('\n', '').strip())
            except Exception as e:
                print('JSON NOT PARSED')
                print(el['name'])
                print(el['appinfo'])

for k, v in enums_xsd_name_dict.items():
    new_dict = dict()
    for k1, v1 in v.items():
        new_dict[k1.lower().strip()] = v1.lower().strip()


def reformat(x):
    x = x.lower().strip()
    x = ' '.join(x.split())
    return x


def reformat_dict(d):
    d = d.copy()
    for k, v in d.items():
        d[k] = reformat(v)
    return d


for enum_name, enum_dict in enums_dict.items():
    if enum_name not in enums_xsd_name_dict:
        print(enum_name, ' NOT FOUND IN XSD')
    elif reformat_dict(enums_xsd_name_dict[enum_name]) != reformat_dict(enum_dict):
        enum_xsd = enums_xsd_name_dict[enum_name]
        print('DIFF BETWEEN ENUMS IN XSD AND MDD')
        print(f'MDD {enum_name}')
        print(f'DIFF WITH XSD {enum_name}')
        for k, v in enum_dict.items():
            if k not in enum_xsd:
                print('missing' + str({k: v}))
            elif reformat(v) != reformat(enum_xsd[k]):
                print({k: (reformat(v), reformat(enum_xsd[k]))})
    elif enums_xsd_name_dict[enum_name] == enum_dict:
        print(f'GOOD MATCH :{enum_name}')

# # edit enum

with open(path / 'deleted_enums.json', 'r', encoding='utf-8') as f:
    deleted_enums = json.load(f)


def apply_restriction_enum(simpletype, name, value_dict):
    restriction = simpletype.find('xs:restriction', namespaces=namespaces)
    if restriction is not None:
        simpletype.remove(restriction)
    restriction = SubElement(simpletype, f"{xs}restriction")
    enum_values = list()
    for el in value_dict[name]:
        try:
            int_el = int(float(el))
            if int_el == float(el):
                enum_values.append(int_el)
            else:
                enum_values.append(str(el))
        except:
            enum_values.append(str(el))

    deleted_enums_values = deleted_enums.get(name, [])
    enum_values = [el for el in enum_values if el not in deleted_enums_values]
    is_int = all([str(el).isdigit() for el in enum_values])

    if is_int:

        enum_values = [int(el) for el in enum_values]
        restriction.attrib.update({"base": 'xs:int'})
        enum_range = range(min(enum_values), max(enum_values) + 1)
        if set(enum_values) != set(enum_range):
            for value in enum_values:
                enumeration = SubElement(restriction, f"{xs}enumeration")
                enumeration.attrib.update({'value': str(value)})
        else:
            min_value = min(enum_values)
            max_value = max(enum_values)
            minInclusive = SubElement(restriction, f"{xs}minInclusive")
            minInclusive.attrib.update({'value': str(min_value)})
            maxInclusive = SubElement(restriction, f"{xs}maxInclusive")
            maxInclusive.attrib.update({'value': str(max_value)})
    else:
        enum_values = [str(el) for el in value_dict[name]]
        restriction.attrib.update({"base": 'xs:string'})
        for value in enum_values:
            enumeration = SubElement(restriction, f"{xs}enumeration")
            enumeration.attrib.update({'value': value})


enum_values = {k: list(v.keys()) for k, v in enums_dict.items()}

for el in enums_xsd:
    enum_name = el.get('name').replace('enum_', '').replace('_id', '')
    self = el['self']
    if enum_name in enums_dict:
        # restrictions
        if self.attrib.get('type', '').startswith('s_'):
            s_type = self.attrib['type']
            s = root.xpath(f'//xs:simpleType[@name="{s_type}"]', namespaces=namespaces)[0]
            apply_restriction_enum(s, enum_name, enum_values)
            appinfo = s.find(f'xs:annotation/xs:appinfo', namespaces=namespaces)
            appinfo.text = '\n' + json.dumps(enums_dict[enum_name], ensure_ascii=False, indent=4) + '\n'
        else:
            if self.attrib.get('type') is not None:
                raise Exception(f'{enum_name} should not have a type')
            else:
                simpletype = self.find('xs:simpleType', namespaces=namespaces)
                if simpletype is not None:
                    self.remove(simpletype)
                simpletype = SubElement(self, f"{xs}simpleType")
                apply_restriction_enum(simpletype, enum_name, enum_values)
        # doc
        appinfo = self.find(f'xs:annotation/xs:appinfo', namespaces=namespaces)
        appinfo.text = '\n' + json.dumps(enums_dict[enum_name], ensure_ascii=False, indent=4) + '\n'
    else:
        raise Exception(f'{enum_name} not found')

special_mapping = {'classe_etiquette': 's_classe_etiquette',
                   'qualite_composant': 's_qualite'}

if WRITE_IN_PLACE is True:
    et.write(str((path / 'modele_commun_DPE_audit.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
else:
    et.write('modele_commun_DPE_audit.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')


## génération des artefacts DPE et audit


audit = copy.deepcopy(root)
dpe = copy.deepcopy(root)
obs_dpe_model_usage = obsdpe + 'model_usage'

## artefacts DPE

et = dpe.getroottree()
audit_to_remove = dpe.find(f'xs:element[@name="audit"]', namespaces=namespaces)
dpe.remove(audit_to_remove)
#
# logement = dpe.find(f'xs:element[@name="logement"]', namespaces=namespaces)
# administratif = dpe.find(f'xs:element[@name="administratif"]', namespaces=namespaces)

# suppression des données audit dans le DPE
for el_audit in list(dpe.iterfind(f'*//xs:element[@obsdpe:model_usage="audit"]', namespaces=namespaces)):
    print(f'removing {el_audit.attrib["name"]} from DPE')
    el_audit.getparent().remove(el_audit)

# for el_audit in administratif.iterfind(f'*//xs:element[@obsdpe:model_usage="audit"]', namespaces=namespaces):
#     el_audit.getparent().remove(el_audit)

# suppression des attributs model_usage

for el in dpe.iterfind('*//xs:element', namespaces=namespaces):

    if obs_dpe_model_usage in el.attrib:
        del el.attrib[obs_dpe_model_usage]

# for el in administratif.iterfind('*//xs:element', namespaces=namespaces):
#
#     if obs_dpe_model_usage in el.attrib:
#         del el.attrib[obs_dpe_model_usage]


## nillable

for el in dpe.xpath('//xs:element[@minOccurs="0"]', namespaces=namespaces):
    el.attrib.update({"nillable": "true"})

# remove annotation modèle commun

dpe.remove(dpe.find('xs:annotation', namespaces=namespaces))

# repassage en relation d'arbre explicite plutôt qu'en référence pour la section logement

logement = dpe.find('xs:element[@name="logement"]', namespaces=namespaces)
logement_copy = copy.deepcopy(logement)
for el in logement_copy.iterfind('*//xs:element', namespaces=namespaces):
    appinfo = el.find('xs:annotation/xs:appinfo', namespaces=namespaces)
    if appinfo is not None:
        appinfo.attrib['source'] = 'dpe/' + appinfo.attrib['source']
dpe.remove(logement)
logement_ref = dpe.find('*//xs:element[@ref="logement"]', namespaces=namespaces)
p = logement_ref.getparent()
p.remove(logement_ref)
p.insert(0, logement_copy)

# repassage en relation d'arbre explicite plutôt qu'en référence pour la section dpe_immeuble

dpe_immeuble = dpe.find('xs:element[@name="dpe_immeuble"]', namespaces=namespaces)
dpe_immeuble_copy = copy.deepcopy(dpe_immeuble)
dpe_immeuble_copy.attrib.update({'minOccurs': '0',"nillable": "true"})
for el in dpe_immeuble_copy.iterfind('*//xs:element', namespaces=namespaces):
    appinfo = el.find('xs:annotation/xs:appinfo', namespaces=namespaces)
    if appinfo is not None:
        appinfo.attrib['source'] = 'dpe/' + appinfo.attrib['source']
dpe.remove(dpe_immeuble)
dpe_immeuble_ref = dpe.find('*//xs:element[@ref="dpe_immeuble"]', namespaces=namespaces)
p = dpe_immeuble_ref.getparent()
p.remove(dpe_immeuble_ref)
p.insert(2, dpe_immeuble_copy)

## write DPE complet

if WRITE_IN_PLACE is True:
    path_dpe_complet = path / 'DPE_complet.xsd'

    et.write(str((path_dpe_complet).absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
else:
    et.write('DPE_complet.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

# ## SET neuf optionnel


logement_neuf = dpe.xpath('//xs:element[@name="logement_neuf"]', namespaces=namespaces)[0]

for el in ['enveloppe', 'ventilation_collection',
           'installation_chauffage_collection',
           'installation_ecs_collection', 'climatisation_collection',
           'production_elec_enr', 'deperdition', 'apport_et_besoin', 'meteo']:
    el = logement_neuf.xpath(f'.//xs:element[@name="{el}"]', namespaces=namespaces)[0]
    el.attrib.update({'minOccurs': '0'})
    for t in el.xpath('.//xs:element', namespaces=namespaces):
        t.attrib.update({'minOccurs': '0'})
for el in dpe.xpath('//xs:element[@minOccurs="0"]', namespaces=namespaces):
    el.attrib.update({"nillable": "true"})

# ## WRITE  neuf optionnel

dpe_v3 = copy.deepcopy(dpe)
et_v3 = dpe_v3.getroottree()

# et_v3_v25 = copy.deepcopy(et_v3)

# suppression de paramètres renommés

# list_remove_parameters = ["invar_logement", "paroi_ancienne"]
#
# for parameter in list_remove_parameters:
#     for el in dpe_v3.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.getparent().remove(el)

CURRENT_VERSION = '2.6'

if WRITE_IN_PLACE is True:
    et_v3.write(str((path / f'DPEv{CURRENT_VERSION}.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
else:
    et_v3.write(f'DPEv{CURRENT_VERSION}.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

#  DPE 2.5

if WRITE_IN_PLACE is True:
    et_v3.write(str((path / f'DPEv2.5.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
else:
    et_v3.write(f'DPEv2.5.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')


# # suppression des données qui disparraissent en 2.5
#
# list_remove_parameters = ["consentement_proprietaire", "information_consentement_proprietaire"]
#
# for parameter in list_remove_parameters:
#     for el in dpe_v3.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.getparent().remove(el)
#
# # données qui deviennent obligatoire en 2.5
#
# list_required_parameters = ["enum_consentement_formulaire_id","enum_commanditaire_id",'paroi_lourde']
#
# for parameter in list_required_parameters:
#     for el in dpe_v3.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         del el.attrib['minOccurs']
#         del el.attrib['nillable']
#
# elements_optionnels = dpe_v3.xpath('//xs:element[@minOccurs="0"]', namespaces=namespaces)
#
# elements = dpe_v3.xpath('//xs:element', namespaces=namespaces)
#
# elements_paths = [(get_sources(el),'obligatoire') if el not in elements_optionnels else (get_sources(el),'optionnel') for el in elements ]
#
# pd.DataFrame(elements_paths,columns=['objet','statut']).to_excel('elements_obligatoires_dpe.xlsx',index=False)
#
# if WRITE_IN_PLACE is True:
#     et_v3.write(str((path / f'DPEv{CURRENT_VERSION}.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et_v3.write(f'DPEv{CURRENT_VERSION}.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')


# #  DPE 2.4
#
# if WRITE_IN_PLACE is True:
#     et_v3_v24.write(str((path / f'DPEv2.4.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et_v3_v24.write(f'DPEv2.4.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')


# DPE 2.3 = DPE 2.4 pour le moment

# if WRITE_IN_PLACE is True:
#     et_v3.write(str((path / f'DPEv2.3.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et_v3.write(f'DPEv2.3.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')


# # DPE 2.2
#
# list_new_parameters = ["label_brut_avec_complement", "rdim", "presence_joint","enduit_isolant_paroi_ancienne","paroi_ancienne","presence_protection_solaire_hors_fermeture"]
#
# for parameter in list_new_parameters:
#     # optionnel
#     for el in dpe.iterfind(f'*//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.attrib.update({'minOccurs': '0'})
#         el.attrib.update({"nillable": "true"})
#
# # optionnel de la surface habitable logement dans dpe_immeuble
# for el in dpe.iterfind(f'*//xs:element[@name="dpe_immeuble"]//xs:element[@name="surface_habitable_logement"]', namespaces=namespaces):
#     el.attrib.update({'minOccurs': '0'})
#     el.attrib.update({"nillable": "true"})
#
#
# # if WRITE_IN_PLACE is True:
# #     et.write(str((path / f'DPEv2.2.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# # else:
# #     et.write(f'DPEv2.2.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')
#
# # DPE 2.1
#
# ## reference en optionnel sur le DPE
#
# for el in dpe.iterfind('*//xs:element[@name="reference"]', namespaces=namespaces):
#     el.attrib.update({'minOccurs': '0'})
#     el.attrib.update({"nillable": "true"})
#
# ## enum_type_pose_id en optionnel pour les portes
#
# for el in dpe.iterfind('*//xs:element[@name="enum_type_pose_id"]', namespaces=namespaces):
#     appinfo = el.find('xs:annotation/xs:appinfo', namespaces=namespaces)
#     if 'porte' in appinfo.attrib['source']:
#         el.attrib.update({'minOccurs': '0'})
#         el.attrib.update({"nillable": "true"})
#
# # pourcentage valeur pont thermique en optionnel
#
# for el in dpe.iterfind('*//xs:element[@name="pourcentage_valeur_pont_thermique"]', namespaces=namespaces):
#     el.attrib.update({'minOccurs': '0'})
#     el.attrib.update({"nillable": "true"})
#
# # restriction string simple pour invar_logement, idpar
#
# invar_logement = dpe.find('*//xs:element[@name="invar_logement"]', namespaces=namespaces)
# st = invar_logement.find('xs:simpleType', namespaces=namespaces)
# invar_logement.remove(st)
# invar_logement.attrib['type'] = "xs:string"
#
# idpar = dpe.find('*//xs:element[@name="idpar"]', namespaces=namespaces)
# st = idpar.find('xs:simpleType', namespaces=namespaces)
# idpar.remove(st)
# idpar.attrib['type'] = "xs:string"

# DPEV2 non maintenu
# if WRITE_IN_PLACE is True:
#     et.write(str((path / 'DPEv2.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et.write('DPEv2.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

## artefact audit

et = audit.getroottree()
dpe_to_remove = audit.find(f'xs:element[@name="dpe"]', namespaces=namespaces)
audit.remove(dpe_to_remove)

# logement = audit.find(f'xs:element[@name="logement"]', namespaces=namespaces)
# administratif = audit.find(f'xs:element[@name="administratif"]', namespaces=namespaces)

# suppression des données DPE dans l'audit
for el_dpe in list(audit.iterfind(f'*//xs:element[@obsdpe:model_usage="dpe"]', namespaces=namespaces)):
    print(f'removing {el_dpe.attrib["name"]} from AUDIT')
    el_dpe.getparent().remove(el_dpe)

# for el_dpe in administratif.iterfind(f'*//xs:element[@obsdpe:model_usage="dpe"]', namespaces=namespaces):
#     el_dpe.getparent().remove(el_dpe)

# suppression des attributs model_usage

for el in audit.iterfind('*//xs:element', namespaces=namespaces):

    if obs_dpe_model_usage in el.attrib:
        del el.attrib[obs_dpe_model_usage]
#
# for el in administratif.iterfind('*//xs:element', namespaces=namespaces):
#
#     if obs_dpe_model_usage in el.attrib:
#         del el.attrib[obs_dpe_model_usage]


# remove annotation modèle commun

audit.remove(audit.find('xs:annotation', namespaces=namespaces))

# repassage en relation d'arbre explicite plutôt qu'en référence pour la section logement

logement = audit.find('xs:element[@name="logement"]', namespaces=namespaces)
logement_copy = copy.deepcopy(logement)
for el in logement_copy.iterfind('*//xs:element', namespaces=namespaces):
    appinfo = el.find('xs:annotation/xs:appinfo', namespaces=namespaces)
    if appinfo is not None:
        appinfo.attrib['source'] = 'audit/' + appinfo.attrib['source']
audit.remove(logement)
logement_ref = audit.find('*//xs:element[@ref="logement"]', namespaces=namespaces)
attrib = copy.copy(logement_ref.attrib)
del attrib['ref']
logement_copy.attrib.update(attrib)
p = logement_ref.getparent()
p.remove(logement_ref)
p.insert(0, logement_copy)


# repassage en relation d'arbre explicite plutôt qu'en référence pour la section dpe_immeuble

dpe_immeuble = audit.find('xs:element[@name="dpe_immeuble"]', namespaces=namespaces)
dpe_immeuble_copy = copy.deepcopy(dpe_immeuble)
dpe_immeuble_copy.attrib.update({'minOccurs': '0',"nillable": "true"})
for el in dpe_immeuble_copy.iterfind('*//xs:element', namespaces=namespaces):
    appinfo = el.find('xs:annotation/xs:appinfo', namespaces=namespaces)
    if appinfo is not None:
        appinfo.attrib['source'] = 'audit/' + appinfo.attrib['source']
audit.remove(dpe_immeuble)
dpe_immeuble_ref = audit.find('*//xs:element[@ref="dpe_immeuble"]', namespaces=namespaces)
p = dpe_immeuble_ref.getparent()
p.remove(dpe_immeuble_ref)
p.insert(2, dpe_immeuble_copy)


# audit = audit permissible
if WRITE_IN_PLACE is True:
    audit.getroottree().write(str((path / 'audit.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
else:
    audit.getroottree().write('audit.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')


audit_v24 = copy.deepcopy(audit)
et_v25 = audit_v24.getroottree()
# et_v24 = copy.deepcopy(et_v25)

CURRENT_VERSION = '2.5'

#  AUDIT 2.5

if WRITE_IN_PLACE is True:
    path_audit_reg = path / 'audit.xsd'
    et_v25.write(str((path / f'audit_v{CURRENT_VERSION}.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')

else:
    et_v25.write(f'audit_v{CURRENT_VERSION}.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

#  AUDIT 2.4

if WRITE_IN_PLACE is True:
    et_v25.write(str((path / f'audit_v2.4.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
else:
    et_v25.write(f'audit_v2.4.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

# suppression de paramètres renommés

# list_remove_parameters = ["invar_logement", "paroi_ancienne"]
#
# for parameter in list_remove_parameters:
#     for el in audit_v21.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.getparent().remove(el)

# suppression des données qui disparraissent en 2.4
#
# list_remove_parameters = ["consentement_proprietaire", "information_consentement_proprietaire"]
#
# for parameter in list_remove_parameters:
#     for el in audit_v24.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.getparent().remove(el)
#
# # données qui deviennent obligatoire en 2.4
#
# list_required_parameters = ["enum_consentement_formulaire_id","enum_commanditaire_id",'paroi_lourde']
#
# for parameter in list_required_parameters:
#     for el in audit_v24.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         del el.attrib['minOccurs']
#         del el.attrib['nillable']
#
# #  AUDIT 2.4
#
# if WRITE_IN_PLACE is True:
#     path_audit_reg = path / 'audit.xsd'
#     et_v24.write(str((path / f'audit_v{CURRENT_VERSION}.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
#
# else:
#     et_v24.write(f'audit_v{CURRENT_VERSION}.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')
#
# #  AUDIT 2.3
#
# if WRITE_IN_PLACE is True:
#     et_v23.write(str((path / f'audit_v2.4.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et_v23.write(f'audit_v2.4.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')
#
# elements_optionnels = et_v24.xpath('//xs:element[@minOccurs="0"]', namespaces=namespaces)
#
# elements = et_v24.xpath('//xs:element', namespaces=namespaces)
#
# elements_paths = [(get_sources(el),'obligatoire') if el not in elements_optionnels else (get_sources(el),'optionnel') for el in elements ]
#
# pd.DataFrame(elements_paths,columns=['objet','statut']).to_excel('elements_obligatoires_audit.xlsx',index=False)


# list_new_parameters = ["label_brut_avec_complement", "rdim", "presence_joint","enduit_isolant_paroi_ancienne","paroi_ancienne","presence_protection_solaire_hors_fermeture"]
#
# for parameter in list_new_parameters:
#     # optionnel
#     for el in audit.iterfind(f'*//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.attrib.update({'minOccurs': '0'})
#         el.attrib.update({"nillable": "true"})
#
# # optionnel de la surface habitable logement dans dpe_immeuble
# for el in audit.iterfind(f'*//xs:element[@name="dpe_immeuble"]//xs:element[@name="surface_habitable_logement"]', namespaces=namespaces):
#     el.attrib.update({'minOccurs': '0'})
#     el.attrib.update({"nillable": "true"})

# # GESTION RETROCOMPATIBILITE audit v2.1
#
# # suppression de paramètres renommés
# list_remove_parameters = ["invar_logement", "paroi_ancienne"] # IL FAUT LAISSER CES DEUX ELEMENTS DANS CETTE LISTE, MEME POUR LA VERSION SUIVANTE. SINON, ces elements seront ajoutés au XSD...
#
# for parameter in list_remove_parameters:
#     for el in audit.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.getparent().remove(el)
#
# list_new_parameters = []
#
# for parameter in list_new_parameters:
#     # optionnel
#     for el in audit.iterfind(f'*//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.attrib.update({'minOccurs': '0'})
#         el.attrib.update({"nillable": "true"})
#
# if WRITE_IN_PLACE is True:
#     path_audit_reg = path / 'audit_v2.1.xsd'
#     et.write(str((path / 'audit_v2.1.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et.write('audit.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

# # GESTION RETROCOMPATIBILITE audit v2.0
#
# # suppression de paramètres renommés
# list_remove_parameters = ["invar_logement", "paroi_ancienne"]
#
# for parameter in list_remove_parameters:
#     for el in audit.xpath(f'//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.getparent().remove(el)
#
# list_new_parameters = ["enum_etat_ventilation_id", "ubat_base", "enum_derogation_ventilation_id"]
#
# for parameter in list_new_parameters:
#     # optionnel
#     for el in audit.iterfind(f'*//xs:element[@name="{parameter}"]', namespaces=namespaces):
#         el.attrib.update({'minOccurs': '0'})
#         el.attrib.update({"nillable": "true"})
#
# if WRITE_IN_PLACE is True:
#     path_audit_reg = path / 'audit_v2.0.xsd'
#     et.write(str((path / 'audit_v2.0.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et.write('audit.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

# GESTION RETROCOMPATIBILITE audit v1.1 ==> Plus besoin, car version plus utilisée depuis septembre 2023
# justificatif_audit_collection = audit.find(f'*//xs:element[@name="justificatif_audit_collection"]', namespaces=namespaces)
# justificatif_audit_collection.attrib.update({'minOccurs': '0'})
# justificatif_audit_collection.attrib.update({"nillable": "true"})
#
# if WRITE_IN_PLACE is True:
#     path_audit_reg = path / 'audit_regv1.1.xsd'
#     et.write(str((path / 'audit_regv1.1.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et.write('audit.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')

# # GESTION RETROCOMPATIBILITE audit v1
#
# for el in audit.iterfind(f'*//xs:element[@name="enum_travaux_resume_id"]', namespaces=namespaces):
#     hard_copy = copy.deepcopy(el)
#     hard_copy.attrib.update({'name':'enum_travaux_resume_collection_id'})
#     el.attrib.update({'minOccurs': '0'})
#     el.attrib.update({"nillable": "true"})
#     hard_copy.attrib.update({'minOccurs': '0'})
#     hard_copy.attrib.update({"nillable": "true"})
#     el.addnext(hard_copy)
#
#
# for el in audit.iterfind(f'*//xs:element[@name="facture_gain"]', namespaces=namespaces):
#     hard_copy_min = copy.deepcopy(el)
#     hard_copy_min.attrib.update({'name': 'facture_gain_min'})
#     hard_copy_min.attrib.update({'minOccurs': '0'})
#     hard_copy_min.attrib.update({"nillable": "true"})
#     hard_copy_max = copy.deepcopy(el)
#     hard_copy_max.attrib.update({'name': 'facture_gain_max'})
#     hard_copy_max.attrib.update({'minOccurs': '0'})
#     hard_copy_max.attrib.update({"nillable": "true"})
#     el.attrib.update({'minOccurs': '0'})
#     el.attrib.update({"nillable": "true"})
#     el.addnext(hard_copy_min)
#     el.addnext(hard_copy_max)
#
# for el in audit.iterfind(f'*//xs:element[@name="facture_gain_cumule"]', namespaces=namespaces):
#     hard_copy_min = copy.deepcopy(el)
#     hard_copy_min.attrib.update({'name': 'facture_gain_cumule_min'})
#     hard_copy_min.attrib.update({'minOccurs': '0'})
#     hard_copy_min.attrib.update({"nillable": "true"})
#     hard_copy_max = copy.deepcopy(el)
#     hard_copy_max.attrib.update({'name': 'facture_gain_cumule_max'})
#     hard_copy_max.attrib.update({'minOccurs': '0'})
#     hard_copy_max.attrib.update({"nillable": "true"})
#     el.attrib.update({'minOccurs': '0'})
#     el.attrib.update({"nillable": "true"})
#     el.addnext(hard_copy_max)
#     el.addnext(hard_copy_min)
#
# if WRITE_IN_PLACE is True:
#     path_audit_reg = path / 'audit_regv1.xsd'
#     et.write(str((path / 'audit_regv1.xsd').absolute()), pretty_print=True, xml_declaration=True, encoding='utf-8')
# else:
#     et.write('audit.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')


# ====================================== EXPORT EXCEL BASE APPLICATIVE ====================================

audit_reg_xsd = etree.parse(str(path_audit_reg), parser)
root_audit = audit_reg_xsd.getroot()

dpe_complet_xsd = etree.parse(str(path_dpe_complet), parser)
root_dpe_complet = dpe_complet_xsd.getroot()

all_el_dpe_complet = list(root_dpe_complet.iterfind('*//xs:element', namespaces=namespaces))
all_el_audit_reg = list(root_audit.iterfind('*//xs:element', namespaces=namespaces))

all_enum_dpe_complet = [el.attrib['name'] for el in all_el_dpe_complet if el.attrib.get('name', 'notaname').startswith('enum_')]

all_enum_audit_reg = [el.attrib['name'] for el in all_el_audit_reg if el.attrib.get('name', 'notaname').startswith('enum_')]

enums_only_dpe = [el.split('enum_')[-1].split('_id')[0] for el in all_enum_dpe_complet if el not in all_enum_audit_reg]
enums_common = [el.split('enum_')[-1].split('_id')[0] for el in all_enum_dpe_complet if el in all_enum_audit_reg]

export_excel_base_applicative_path = path / 'base_applicative' / "export_excel_base_applicative"
export_excel_base_applicative_path.mkdir(exist_ok=True, parents=True)

enum_tables_only_dpe = openpyxl.load_workbook(path / 'enum_tables.xlsx')

# del all sheets that are not exclusive to dpe
to_del = [el for el in enum_tables_only_dpe.sheetnames if el not in enums_only_dpe + ['index']]
for el in to_del:
    del enum_tables_only_dpe[el]

enum_tables_only_dpe.save(export_excel_base_applicative_path / 'enum_tables_dpe_only.xlsx')

enum_tables_common = openpyxl.load_workbook(path / 'enum_tables.xlsx')

# del all sheets that are not exclusive to dpe
to_del = [el for el in enum_tables_common.sheetnames if el not in enums_common + ['index', 'classe_etiquette', "qualite_composant"]]
for el in to_del:
    del enum_tables_common[el]

enum_tables = pd.read_excel(path / 'enum_tables.xlsx', sheet_name=None)

all_enums_dpe = enum_tables.keys()

all_enums_dpe_rebuild = set(enum_tables_common.sheetnames) | set(enum_tables_only_dpe.sheetnames)

diff = set(all_enums_dpe) - set(all_enums_dpe_rebuild)
if len(diff) > 0:
    raise KeyError(f'missing {diff} in split enum tables')

enum_tables_common.save(export_excel_base_applicative_path / 'enum_tables_common.xlsx')

enum_tables_audit = openpyxl.load_workbook(path / 'enum_tables_audit.xlsx')

enum_tables_audit.save(export_excel_base_applicative_path / 'enum_tables_audit.xlsx')

# =============================== GENERATION d'un XSD completement optionnel =============================
et = dpe.getroottree()
for el in dpe.xpath('//xs:element', namespaces=namespaces):
    if not 'ref' in el.attrib:
        el.attrib.update({'minOccurs': "0", "nillable": "true"})
et.write('DPE_tout_optionnel.xsd', pretty_print=True, xml_declaration=True, encoding='utf-8')
